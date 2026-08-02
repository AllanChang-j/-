from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import shutil
import ssl
import urllib.error
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import openpyxl

try:
    import certifi
except ImportError:  # pragma: no cover - fallback for environments without certifi
    certifi = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG = PROJECT_ROOT / "config" / "sources.json"
RAW_SHEET = "1收盤"
RAW_START_ROW = 8
RAW_END_ROW = 2349
RAW_START_COL = 1
RAW_END_COL = 17
FORMULA_START_COL = 20
FORMULA_END_COL = 70


LISTED_COLUMNS = [
    "證券代號",
    "證券名稱",
    "成交股數",
    "成交筆數",
    "成交金額",
    "開盤價",
    "最高價",
    "最低價",
    "收盤價",
    "漲跌(+/-)",
    "漲跌價差",
    "最後揭示買價",
    "最後揭示買量",
    "最後揭示賣價",
    "最後揭示賣量",
    "本益比",
]

MAINBOARD_COLUMNS = [
    "代號",
    "名稱",
    "收盤",
    "漲跌",
    "開盤",
    "最高",
    "最低",
    "成交股數",
    "成交金額",
    "成交筆數",
    "最後買價",
    "最後買量(張)",
    "最後賣價",
    "最後賣量(張)",
    "發行股數",
    "次日漲停價",
    "次日跌停價",
]

ESB_COLUMNS = [
    "代號",
    "名稱",
    "前日均價",
    "報買價",
    "報買量",
    "報賣價",
    "報賣量",
    "日最高",
    "日最低",
    "日均價",
    "成交",
    "投資人成交買賣別",
    "暫停交易開始時間",
    "成交量",
    "進度日期",
    "上市櫃進度",
]

TEXT_COLUMNS_BY_MARKET = {
    "listed": {1, 2, 10},
    "mainboard": {1, 2},
    "esb": {1, 2, 12, 13, 15, 16},
}

MARKET_LABELS = {
    "listed": "上市",
    "mainboard": "上櫃",
    "esb": "興櫃",
}

MARKET_ORDER = {
    "listed": 0,
    "mainboard": 1,
    "esb": 2,
}

VOLUME_COLUMN_BY_MARKET = {
    "listed": 3,
    "mainboard": 8,
    "esb": 14,
}

AMOUNT_COLUMN_BY_MARKET = {
    "listed": 5,
    "mainboard": 9,
    "esb": None,
}


@dataclass(frozen=True)
class SourceConfig:
    key: str
    market: str
    label: str
    url: str
    fmt: str
    method: str = "GET"
    body: dict[str, str] | None = None
    historical: bool = True
    allow_insecure_ssl_fallback: bool = False


@dataclass(frozen=True)
class PreparedRow:
    market: str
    label: str
    symbol: str
    name: str
    values: list[Any]
    from_source: bool = True


@dataclass(frozen=True)
class TemplateRawRow:
    row_index: int
    market: str
    symbol: str
    name: str
    values: list[Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="更新既有收盤日報 Excel 範本。")
    parser.add_argument("--template", default=os.environ.get("STAGE1_TEMPLATE"), help="原始 xlsx 範本路徑；也可用 STAGE1_TEMPLATE")
    parser.add_argument("--date", help="交易日，格式 YYYY-MM-DD；省略時使用今天")
    parser.add_argument("--config", default=os.environ.get("STAGE1_CONFIG", str(DEFAULT_CONFIG)), help="資料來源設定檔")
    parser.add_argument("--output-dir", default=os.environ.get("STAGE1_OUTPUT_DIR", str(PROJECT_ROOT / "outputs")), help="輸出資料夾")
    parser.add_argument("--dry-run", action="store_true", help="只下載與檢查資料，不輸出 xlsx")
    parser.add_argument("--include-non-stock", action="store_true", help="保留非四碼股票代號商品")
    args = parser.parse_args()
    if not args.template:
        parser.error("--template 或 STAGE1_TEMPLATE 必須提供原始 xlsx 範本路徑")
    return args


def roc_date_slash(day: dt.date) -> str:
    return f"{day.year - 1911}/{day.month:02d}/{day.day:02d}"


def load_sources(config_path: Path, day: dt.date) -> list[SourceConfig]:
    config = json.loads(config_path.read_text(encoding="utf-8"))
    sources: list[SourceConfig] = []
    for key, raw in config["sources"].items():
        url = raw["url"].format(
            date_yyyymmdd=day.strftime("%Y%m%d"),
            date_iso=day.isoformat(),
            date_roc_slash=roc_date_slash(day),
        )
        body = raw.get("body")
        if body:
            body = {
                key: value.format(
                    date_yyyymmdd=day.strftime("%Y%m%d"),
                    date_iso=day.isoformat(),
                    date_roc_slash=roc_date_slash(day),
                )
                for key, value in body.items()
            }
        sources.append(
            SourceConfig(
                key=key,
                market=raw["market"],
                label=raw["label"],
                url=url,
                fmt=raw["format"],
                method=raw.get("method", "GET").upper(),
                body=body,
                historical=raw.get("historical", True),
                allow_insecure_ssl_fallback=raw.get("allow_insecure_ssl_fallback", False),
            )
        )
    return sources


def fetch_text(url: str) -> str:
    return fetch_text_with_method(url, "GET", None, False)


def fetch_text_with_method(url: str, method: str, body: dict[str, str] | None, allow_insecure_ssl_fallback: bool) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0 stock-dashboard-stage1/0.1",
        "Accept": "application/json,text/csv,text/plain,*/*",
    }
    data = None
    if method == "POST":
        data = urlencode(body or {}).encode("utf-8")
        headers["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"
    request = Request(url, data=data, headers=headers, method=method)
    context = ssl.create_default_context(cafile=certifi.where()) if certifi else None
    try:
        with urlopen(request, timeout=40, context=context) as response:
            raw = response.read()
            charset = response.headers.get_content_charset() or "utf-8"
    except urllib.error.URLError as exc:
        reason = exc.reason if hasattr(exc, "reason") else exc
        is_cert_error = isinstance(reason, ssl.SSLCertVerificationError)
        if not (allow_insecure_ssl_fallback and is_cert_error):
            raise
        insecure_context = ssl._create_unverified_context()
        with urlopen(request, timeout=40, context=insecure_context) as response:
            raw = response.read()
            charset = response.headers.get_content_charset() or "utf-8"
    try:
        return raw.decode(charset)
    except UnicodeDecodeError:
        return raw.decode("big5", errors="replace")


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"<[^>]+>", "", text)
    return re.sub(r"[\s\u3000]+", "", text.replace("\ufeff", ""))


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    text = re.sub(r"<[^>]+>", "", str(value)).strip()
    if text in {"", "--", "---"}:
        return text
    text = text.replace(",", "")
    return text


def coerce_number(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if text in {"", "--", "---", "-"}:
        return None
    if not re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text):
        return value
    number = float(text)
    return int(number) if number.is_integer() else number


def coerce_template_row(market: str, values: list[Any]) -> list[Any]:
    text_columns = TEXT_COLUMNS_BY_MARKET[market]
    typed_values: list[Any] = []
    for index, value in enumerate(values, start=1):
        if index in text_columns:
            typed_values.append(None if value is None else str(value).strip())
        else:
            typed_values.append(coerce_number(value))
    return typed_values


def find_twse_rows(payload: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    candidates: list[tuple[int, list[str], list[list[Any]]]] = []
    for table in payload.get("tables", []):
        if not isinstance(table, dict):
            continue
        fields = table.get("fields", [])
        data = table.get("data", [])
        joined = "".join(normalize_header(x) for x in fields)
        if "證券代號" in joined and "證券名稱" in joined and isinstance(data, list):
            candidates.append((len(data), fields, data))
    for key, data in payload.items():
        if not key.startswith("data") or not isinstance(data, list):
            continue
        suffix = key[4:]
        fields = payload.get(f"fields{suffix}", [])
        if not fields:
            continue
        joined = "".join(normalize_header(x) for x in fields)
        if "證券代號" in joined and "證券名稱" in joined:
            candidates.append((len(data), fields, data))
    if not candidates:
        raise ValueError("TWSE 回應中找不到每日個股收盤資料表。")
    return max(candidates, key=lambda item: item[0])[1:]


def parse_json_rows(source: SourceConfig, text: str) -> list[dict[str, Any]]:
    payload = json.loads(text)
    if isinstance(payload, list):
        return [{normalize_header(k): clean_cell(v) for k, v in row.items()} for row in payload if isinstance(row, dict)]

    if source.market == "listed":
        fields, rows = find_twse_rows(payload)
        headers = [normalize_header(x) for x in fields]
        return [dict(zip(headers, (clean_cell(x) for x in row))) for row in rows]

    tables = payload.get("tables")
    if isinstance(tables, list):
        candidates = []
        for table in tables:
            if not isinstance(table, dict):
                continue
            fields = table.get("fields", [])
            data = table.get("data", [])
            joined = "".join(normalize_header(x) for x in fields)
            if ("代號" in joined or "SecuritiesCompanyCode" in joined) and isinstance(data, list):
                candidates.append((len(data), fields, data))
        if candidates:
            _, fields, data = max(candidates, key=lambda item: item[0])
            headers = [normalize_header(x) for x in fields]
            return [dict(zip(headers, (clean_cell(x) for x in row))) for row in data]

    for key in ("data", "aaData", "tables"):
        data = payload.get(key)
        if isinstance(data, list) and data:
            if isinstance(data[0], dict):
                return [{normalize_header(k): clean_cell(v) for k, v in row.items()} for row in data]
            if isinstance(data[0], list):
                fields = payload.get("fields") or payload.get("columns")
                if not fields:
                    raise ValueError(f"{source.label} JSON 是陣列資料，但沒有欄名。")
                headers = [normalize_header(x) for x in fields]
                return [dict(zip(headers, (clean_cell(x) for x in row))) for row in data]
    raise ValueError(f"{source.label} JSON 格式無法辨識。")


def parse_csv_rows(text: str) -> list[dict[str, Any]]:
    sample = text.replace("\ufeff", "")
    reader = csv.reader(StringIO(sample))
    rows = list(reader)
    header_index = None
    for index, row in enumerate(rows):
        joined = "".join(normalize_header(x) for x in row)
        if "代號" in joined and "名稱" in joined:
            header_index = index
            break
    if header_index is None:
        raise ValueError("CSV 中找不到含代號與名稱的標題列。")
    headers = [normalize_header(x) for x in rows[header_index]]
    return [
        dict(zip(headers, (clean_cell(x) for x in row)))
        for row in rows[header_index + 1 :]
        if any(str(x).strip() for x in row)
    ]


def source_rows(source: SourceConfig) -> list[dict[str, Any]]:
    try:
        text = fetch_text_with_method(source.url, source.method, source.body, source.allow_insecure_ssl_fallback)
    except urllib.error.URLError as exc:
        detail = exc.reason if hasattr(exc, "reason") else exc
        raise RuntimeError(
            f"{source.label} 資料來源連線失敗：{source.url}\n"
            f"請先確認網路/DNS 可連線，或用瀏覽器打開來源頁確認網站是否正常。\n"
            f"原始錯誤：{detail}"
        ) from exc
    if source.fmt.lower() == "csv":
        return parse_csv_rows(text)
    try:
        return parse_json_rows(source, text)
    except ValueError as exc:
        raise RuntimeError(
            f"{source.label} 資料格式無法辨識或資料尚未發布：{source.url}\n"
            f"若查詢今天，可能是交易所尚未提供當日收盤表；可稍晚再試或指定已發布日期。\n"
            f"原始錯誤：{exc}"
        ) from exc


def is_today(day: dt.date) -> bool:
    return day == dt.date.today()


def row_symbol(row: dict[str, Any]) -> str:
    value = pick(row, "證券代號", "代號", "SecuritiesCompanyCode")
    return "" if value is None else str(value).strip()


def filter_stock_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if re.fullmatch(r"\d{4}", row_symbol(row)) and not row_symbol(row).startswith("00")]


def pick(row: dict[str, Any], *names: str) -> Any:
    normalized = {normalize_header(k): v for k, v in row.items()}
    for name in names:
        key = normalize_header(name)
        if key in normalized:
            return clean_cell(normalized[key])
    return None


def to_template_row(market: str, row: dict[str, Any]) -> list[Any]:
    if market == "listed":
        return coerce_template_row(market, [
            pick(row, "證券代號"),
            pick(row, "證券名稱"),
            pick(row, "成交股數"),
            pick(row, "成交筆數"),
            pick(row, "成交金額"),
            pick(row, "開盤價"),
            pick(row, "最高價"),
            pick(row, "最低價"),
            pick(row, "收盤價"),
            pick(row, "漲跌(+/-)"),
            pick(row, "漲跌價差"),
            pick(row, "最後揭示買價"),
            pick(row, "最後揭示買量"),
            pick(row, "最後揭示賣價"),
            pick(row, "最後揭示賣量"),
            pick(row, "本益比"),
        ])
    if market == "mainboard":
        return coerce_template_row(market, [
            pick(row, "代號", "SecuritiesCompanyCode"),
            pick(row, "名稱", "CompanyName"),
            pick(row, "收盤", "Close"),
            pick(row, "漲跌", "Change"),
            pick(row, "開盤", "Open"),
            pick(row, "最高", "High"),
            pick(row, "最低", "Low"),
            pick(row, "成交股數", "TradingShares"),
            pick(row, "成交金額", "成交金額(元)", "TransactionAmount"),
            pick(row, "成交筆數", "TransactionNumber"),
            pick(row, "最後買價", "LatestBidPrice"),
            pick(row, "最後買量(張)", "最後買量(張數)", "LatestBidQuantity"),
            pick(row, "最後賣價", "LatesAskPrice", "LatestAskPrice"),
            pick(row, "最後賣量(張)", "最後賣量(張數)", "LatestAskQuantity"),
            pick(row, "發行股數", "Capitals"),
            pick(row, "次日漲停價", "NextLimitUp"),
            pick(row, "次日跌停價", "NextLimitDown"),
        ])
    if market == "esb":
        return coerce_template_row(market, [
            pick(row, "代號", "SecuritiesCompanyCode"),
            pick(row, "名稱", "CompanyName"),
            pick(row, "前日均價", "PreviousAveragePrice"),
            pick(row, "報買價", "BuyingPrice"),
            pick(row, "報買量", "BuyingQuantity"),
            pick(row, "報賣價", "SellingPrice"),
            pick(row, "報賣量", "SellingQuantity"),
            pick(row, "日最高", "Highest"),
            pick(row, "日最低", "Lowest"),
            pick(row, "日均價", "Average"),
            pick(row, "成交", "LatestPrice"),
            pick(row, "投資人成交買賣別", "Buy/Sell"),
            pick(row, "暫停交易開始時間", "SuspendTime"),
            pick(row, "成交量", "TransactionVolume"),
            pick(row, "進度日期", "ApplyingDate"),
            pick(row, "上市櫃進度", "ApplyingStatus"),
        ])
    raise ValueError(f"未知市場：{market}")


def clear_raw_row(ws: openpyxl.worksheet.worksheet.Worksheet, row_index: int) -> None:
    for col_index in range(RAW_START_COL, RAW_END_COL + 1):
        ws.cell(row_index, col_index).value = None


def fill_down_formulas(ws: openpyxl.worksheet.worksheet.Worksheet, from_row: int, to_row: int) -> None:
    if to_row <= from_row:
        return
    for col in range(FORMULA_START_COL, FORMULA_END_COL + 1):
        source_cell = ws.cell(from_row, col)
        if not isinstance(source_cell.value, str) or not source_cell.value.startswith("="):
            continue
        for row in range(from_row + 1, to_row + 1):
            target = ws.cell(row, col)
            if target.value is None or isinstance(target.value, str) and target.value.startswith("="):
                target.value = openpyxl.formula.translate.Translator(source_cell.value, origin=source_cell.coordinate).translate_formula(target.coordinate)


def formula_snapshot(wb: openpyxl.Workbook) -> dict[tuple[str, str], str]:
    formulas: dict[tuple[str, str], str] = {}
    for ws in wb.worksheets:
        for cell in ws._cells.values():
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas[(ws.title, cell.coordinate)] = cell.value
    return formulas


def ensure_raw_cells_loaded(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for row_index in range(RAW_START_ROW, RAW_END_ROW + 1):
        for col_index in range(RAW_START_COL, RAW_END_COL + 1):
            ws.cell(row_index, col_index)


def style_snapshot(wb: openpyxl.Workbook) -> dict[tuple[str, str], int]:
    styles: dict[tuple[str, str], int] = {}
    for ws in wb.worksheets:
        for cell in ws._cells.values():
            styles[(ws.title, cell.coordinate)] = cell.style_id
    return styles


def sheet_format_snapshot(wb: openpyxl.Workbook) -> dict[str, dict[str, Any]]:
    snapshot: dict[str, dict[str, Any]] = {}
    for ws in wb.worksheets:
        snapshot[ws.title] = {
            "freeze_panes": str(ws.freeze_panes or ""),
            "auto_filter": str(ws.auto_filter.ref or ""),
            "merged_ranges": tuple(sorted(str(item) for item in ws.merged_cells.ranges)),
            "show_gridlines": bool(ws.sheet_view.showGridLines),
            "row_dimensions": tuple(
                sorted(
                    (
                        index,
                        dimension.height,
                        dimension.hidden,
                        dimension.outlineLevel,
                        dimension.style,
                    )
                    for index, dimension in ws.row_dimensions.items()
                )
            ),
            "column_dimensions": tuple(
                sorted(
                    (
                        index,
                        dimension.width,
                        dimension.hidden,
                        dimension.outlineLevel,
                        dimension.style,
                    )
                    for index, dimension in ws.column_dimensions.items()
                )
            ),
        }
    return snapshot


def validate_styles_unchanged(before: dict[tuple[str, str], int], wb: openpyxl.Workbook) -> None:
    after = style_snapshot(wb)
    changed = []
    for key, style_id in before.items():
        if after.get(key) != style_id:
            changed.append((key, style_id, after.get(key)))
    added_with_style = [
        (key, after[key])
        for key in sorted(set(after) - set(before))
        if after[key] != 0
    ]
    if changed or added_with_style:
        details = []
        for (sheet, cell), old, new in changed[:10]:
            details.append(f"{sheet}!{cell}: style {old!r} -> {new!r}")
        for (sheet, cell), style_id in added_with_style[:10]:
            details.append(f"{sheet}!{cell}: added styled cell {style_id!r}")
        raise ValueError("格式完整性檢查失敗；儲存格樣式不應被自動化流程修改。\n" + "\n".join(details))
    print(f"儲存格格式完整性: OK ({len(before)} styled cells preserved)")


def validate_sheet_formats_unchanged(before: dict[str, dict[str, Any]], wb: openpyxl.Workbook) -> None:
    after = sheet_format_snapshot(wb)
    changed = []
    for sheet_name, before_values in before.items():
        after_values = after.get(sheet_name)
        if after_values != before_values:
            changed.append(sheet_name)
    added = sorted(set(after) - set(before))
    if changed or added:
        details = changed[:10] + [f"{sheet_name}: added sheet" for sheet_name in added[:10]]
        raise ValueError("工作表格式完整性檢查失敗；合併格、凍結窗格、篩選或欄列格式被修改。\n" + "\n".join(details))
    print(f"工作表格式完整性: OK ({len(before)} sheets preserved)")


def validate_formulas_unchanged(before: dict[tuple[str, str], str], wb: openpyxl.Workbook) -> None:
    after = formula_snapshot(wb)
    changed = []
    for key, formula in before.items():
        if after.get(key) != formula:
            changed.append((key, formula, after.get(key)))
    added = sorted(set(after) - set(before))
    if changed or added:
        details = []
        for (sheet, cell), old, new in changed[:10]:
            details.append(f"{sheet}!{cell}: {old!r} -> {new!r}")
        for sheet, cell in added[:10]:
            details.append(f"{sheet}!{cell}: added {after[(sheet, cell)]!r}")
        raise ValueError("公式完整性檢查失敗；公式區不應被自動化流程修改。\n" + "\n".join(details))
    print(f"公式完整性: OK ({len(before)} formulas preserved)")


def raw_row_values(ws: openpyxl.worksheet.worksheet.Worksheet, row_index: int) -> list[Any]:
    return [ws.cell(row_index, col_index).value for col_index in range(RAW_START_COL, RAW_END_COL + 1)]


def normalize_symbol(value: Any) -> str:
    return "" if value is None else str(value).strip()


def template_raw_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[TemplateRawRow]:
    rows: list[TemplateRawRow] = []
    current_market = "listed"
    for row_index in range(RAW_START_ROW, RAW_END_ROW + 1):
        values = raw_row_values(ws, row_index)
        symbol = normalize_symbol(values[0])
        name = "" if values[1] is None else str(values[1]).strip()
        if name.startswith("上市資料"):
            current_market = "mainboard"
            continue
        if name.startswith("上櫃資料"):
            current_market = "esb"
            continue
        if re.fullmatch(r"\d{4}", symbol):
            rows.append(TemplateRawRow(row_index, current_market, symbol, name, values))
    return rows


def prepare_source_rows(rows_by_source: list[tuple[SourceConfig, list[dict[str, Any]]]]) -> list[PreparedRow]:
    prepared: list[PreparedRow] = []
    for source, rows in rows_by_source:
        for row in rows:
            values = to_template_row(source.market, row)
            symbol = normalize_symbol(values[0])
            if not symbol:
                continue
            name = "" if values[1] is None else str(values[1]).strip()
            prepared.append(PreparedRow(source.market, source.label, symbol, name, values))
    return prepared


def preserved_template_rows(template_rows: list[TemplateRawRow], fetched_markets: set[str]) -> list[PreparedRow]:
    preserved: list[PreparedRow] = []
    for row in template_rows:
        if row.market in fetched_markets:
            continue
        preserved.append(PreparedRow(row.market, MARKET_LABELS.get(row.market, row.market), row.symbol, row.name, row.values, False))
    return preserved


def market_sort_key(row: PreparedRow) -> tuple[int, str]:
    return (MARKET_ORDER.get(row.market, 99), row.symbol)


def compare_cell_value(expected: Any, actual: Any) -> bool:
    if expected is None and actual in (None, ""):
        return True
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return abs(float(expected) - float(actual)) < 0.0000001
    return expected == actual


def validate_written_source_values(ws: openpyxl.worksheet.worksheet.Worksheet, prepared_rows: list[PreparedRow]) -> dict[str, dict[str, Any]]:
    errors: list[str] = []
    summary: dict[str, dict[str, Any]] = {}
    numeric_cells = 0
    source_rows = 0
    preserved_rows = 0
    for offset, prepared in enumerate(prepared_rows):
        row_index = RAW_START_ROW + offset
        market_summary = summary.setdefault(prepared.market, {"label": prepared.label, "rows": 0, "volume": 0, "amount": 0})
        market_summary["rows"] += 1
        if prepared.from_source:
            source_rows += 1
        else:
            preserved_rows += 1
        for col_offset, expected in enumerate(prepared.values, start=RAW_START_COL):
            actual = ws.cell(row_index, col_offset).value
            if isinstance(expected, (int, float)):
                numeric_cells += 1
            if prepared.from_source and not compare_cell_value(expected, actual):
                errors.append(f"{ws.title}!{ws.cell(row_index, col_offset).coordinate}: source={expected!r}, workbook={actual!r}")
        volume_col = VOLUME_COLUMN_BY_MARKET.get(prepared.market)
        if volume_col and isinstance(prepared.values[volume_col - 1], (int, float)):
            market_summary["volume"] += prepared.values[volume_col - 1]
        amount_col = AMOUNT_COLUMN_BY_MARKET.get(prepared.market)
        if amount_col and isinstance(prepared.values[amount_col - 1], (int, float)):
            market_summary["amount"] += prepared.values[amount_col - 1]
    if errors:
        raise ValueError("數值交叉檢查失敗；來源資料與寫入 Excel 的 raw cell 不一致。\n" + "\n".join(errors[:20]))
    preserved_text = f", preserved template rows={preserved_rows}" if preserved_rows else ""
    print(f"數值交叉檢查: OK (source rows={source_rows}{preserved_text}, numeric cells={numeric_cells})")
    for market in sorted(summary, key=lambda key: MARKET_ORDER.get(key, 99)):
        item = summary[market]
        amount_text = f", amount={int(item['amount'])}" if item["amount"] else ""
        print(f"  {item['label']}: rows={item['rows']}, volume={int(item['volume'])}{amount_text}")
    return summary


def print_symbol_change_report(template_rows: list[TemplateRawRow], prepared_rows: list[PreparedRow], fetched_markets: set[str]) -> None:
    template_by_market: dict[str, dict[str, TemplateRawRow]] = {}
    source_by_market: dict[str, dict[str, PreparedRow]] = {}
    for row in template_rows:
        template_by_market.setdefault(row.market, {})[row.symbol] = row
    for row in prepared_rows:
        if row.from_source:
            source_by_market.setdefault(row.market, {})[row.symbol] = row
    for market in sorted(fetched_markets, key=lambda key: MARKET_ORDER.get(key, 99)):
        template_symbols = set(template_by_market.get(market, {}))
        source_symbols = set(source_by_market.get(market, {}))
        added = sorted(source_symbols - template_symbols)
        removed = sorted(template_symbols - source_symbols)
        label = MARKET_LABELS.get(market, market)
        if not added and not removed:
            print(f"{label}公司清單變化: 無")
            continue
        print(f"{label}公司清單變化: added={len(added)}, removed={len(removed)}")
        if added:
            examples = ", ".join(f"{symbol} {source_by_market[market][symbol].name}".strip() for symbol in added[:10])
            print(f"  新增: {examples}")
        if removed:
            examples = ", ".join(f"{symbol} {template_by_market[market][symbol].name}".strip() for symbol in removed[:10])
            print(f"  刪除: {examples}")


def update_workbook(template: Path, output_path: Path, day: dt.date, rows_by_source: list[tuple[SourceConfig, list[dict[str, Any]]]]) -> None:
    shutil.copy2(template, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb[RAW_SHEET]
    ensure_raw_cells_loaded(ws)
    formulas_before = formula_snapshot(wb)
    styles_before = style_snapshot(wb)
    sheet_formats_before = sheet_format_snapshot(wb)
    template_rows = template_raw_rows(ws)
    fetched_markets = {source.market for source, _ in rows_by_source}
    prepared_rows = prepare_source_rows(rows_by_source)
    preserved_rows = preserved_template_rows(template_rows, fetched_markets)
    combined_rows = sorted(prepared_rows, key=market_sort_key) + sorted(preserved_rows, key=market_sort_key)
    capacity = RAW_END_ROW - RAW_START_ROW + 1
    if len(combined_rows) > capacity:
        raise ValueError(f"來源資料筆數 {len(combined_rows)} 超過範本 raw 區容量 {capacity}，需要擴充範本公式列。")
    ws["A5"] = f"{day.year - 1911}/{day.month}/{day.day}"

    for row_index in range(RAW_START_ROW, RAW_END_ROW + 1):
        clear_raw_row(ws, row_index)

    for offset, prepared in enumerate(combined_rows):
        row_index = RAW_START_ROW + offset
        for col_offset, value in enumerate(prepared.values, start=RAW_START_COL):
            ws.cell(row_index, col_offset).value = value

    validate_written_source_values(ws, combined_rows)
    validate_formulas_unchanged(formulas_before, wb)
    validate_styles_unchanged(styles_before, wb)
    validate_sheet_formats_unchanged(sheet_formats_before, wb)
    wb.save(output_path)
    print_symbol_change_report(template_rows, prepared_rows, fetched_markets)
    print(f"每日來源資料寫入: {len(prepared_rows)} rows")
    if preserved_rows:
        preserved_labels = sorted({row.label for row in preserved_rows})
        print(f"保留未抓取市場範本資料: {len(preserved_rows)} rows ({', '.join(preserved_labels)})")


def main() -> None:
    args = parse_args()
    day = dt.date.fromisoformat(args.date) if args.date else dt.date.today()
    sources = load_sources(Path(args.config), day)

    rows_by_source: list[tuple[SourceConfig, list[dict[str, Any]]]] = []
    for source in sources:
        if not source.historical and not is_today(day):
            print(f"{source.label}: skipped (此來源只支援最新日，指定日期 {day.isoformat()} 不覆蓋範本)")
            continue
        rows = source_rows(source)
        raw_count = len(rows)
        if not args.include_non_stock:
            rows = filter_stock_rows(rows)
        rows_by_source.append((source, rows))
        print(f"{source.label}: {len(rows)} rows" + (f" (raw {raw_count})" if raw_count != len(rows) else ""))

    total = sum(len(rows) for _, rows in rows_by_source)
    print(f"合計來源股票: {total} rows")

    if args.dry_run:
        return

    output_dir = Path(args.output_dir) / day.strftime("%Y%m%d")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"收盤日報_{day.strftime('%Y%m%d')}.xlsx"
    update_workbook(Path(args.template), output_path, day, rows_by_source)
    print(output_path)


if __name__ == "__main__":
    main()
