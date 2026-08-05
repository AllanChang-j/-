from __future__ import annotations

import argparse
import datetime as dt
import json
import time
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from stage1_close_report import (
    DEFAULT_CONFIG,
    filter_stock_rows,
    load_sources,
    source_rows,
    to_template_row,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_HISTORY_CACHE = PROJECT_ROOT / "data" / "raw" / "taiwan_daily_ohlcv_stage1.csv"
EPS = 1e-12

BASE_COLUMNS = [
    "date",
    "market",
    "symbol",
    "name",
    "open",
    "high",
    "low",
    "close",
    "volume",
    "amount",
]

EXPORT_COLUMNS = {
    "date": "日期",
    "market": "市場",
    "symbol": "代號",
    "name": "名稱",
    "open": "開盤",
    "high": "最高",
    "low": "最低",
    "close": "收盤",
    "volume": "今日量",
    "amount": "成交金額",
    "high_20d": "20日最高",
    "low_20d": "20日最低",
    "dist_to_20d_high_pct": "距20日高(%)",
    "dist_to_20d_low_pct": "距20日低(%)",
    "high_60d": "60日最高",
    "low_60d": "60日最低",
    "dist_to_60d_high_pct": "距60日高(%)",
    "dist_to_60d_low_pct": "距60日低(%)",
    "high_120d": "120日最高",
    "low_120d": "120日最低",
    "dist_to_120d_high_pct": "距120日高(%)",
    "dist_to_120d_low_pct": "距120日低(%)",
    "high_240d": "一年最高",
    "low_240d": "一年最低",
    "dist_to_240d_high_pct": "距一年高(%)",
    "dist_to_240d_low_pct": "距一年低(%)",
    "new_high_20d": "創20日新高",
    "new_high_60d": "創60日新高",
    "new_high_240d": "創一年新高",
    "ma_5": "5MA",
    "ma_10": "10MA",
    "ma_20": "20MA",
    "ma_60": "60MA",
    "ma_120": "120MA",
    "ma_240": "240MA",
    "close_to_ma_5_pct": "收盤離5MA(%)",
    "close_to_ma_10_pct": "收盤離10MA(%)",
    "close_to_ma_20_pct": "收盤離20MA(%)",
    "close_to_ma_60_pct": "收盤離60MA(%)",
    "close_to_ma_120_pct": "收盤離120MA(%)",
    "close_to_ma_240_pct": "收盤離240MA(%)",
    "volume_ma_5": "5日均量",
    "volume_ma_20": "20日均量",
    "volume_ma_60": "60日均量",
    "volume_ratio_20": "量比20MA",
    "volume_status": "量能分類",
}

NUMERIC_EXPORT_COLUMNS = {
    "開盤",
    "最高",
    "最低",
    "收盤",
    "今日量",
    "成交金額",
    "20日最高",
    "20日最低",
    "距20日高(%)",
    "距20日低(%)",
    "60日最高",
    "60日最低",
    "距60日高(%)",
    "距60日低(%)",
    "120日最高",
    "120日最低",
    "距120日高(%)",
    "距120日低(%)",
    "一年最高",
    "一年最低",
    "距一年高(%)",
    "距一年低(%)",
    "5MA",
    "10MA",
    "20MA",
    "60MA",
    "120MA",
    "240MA",
    "收盤離5MA(%)",
    "收盤離10MA(%)",
    "收盤離20MA(%)",
    "收盤離60MA(%)",
    "收盤離120MA(%)",
    "收盤離240MA(%)",
    "5日均量",
    "20日均量",
    "60日均量",
    "量比20MA",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="產出每日收盤高低價、均價、均量指標。")
    parser.add_argument("--start", required=True, help="開始日期 YYYY-MM-DD")
    parser.add_argument("--end", required=True, help="結束日期 YYYY-MM-DD")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="Stage 1 資料來源設定")
    parser.add_argument("--history-cache", default=str(DEFAULT_HISTORY_CACHE), help="歷史 OHLCV 快取 CSV")
    parser.add_argument("--seed-cache", help="既有歷史 OHLCV CSV；history-cache 不存在時可用它啟動")
    parser.add_argument("--output-dir", default=str(PROJECT_ROOT / "outputs" / "indicators"), help="指標輸出根目錄")
    parser.add_argument("--request-delay", type=float, default=0.25, help="每個來源請求間隔秒數")
    parser.add_argument("--include-esb-latest", action="store_true", help="只在結束日為今天時納入興櫃最新日資料")
    parser.add_argument("--include-non-stock", action="store_true", help="保留非四碼股票代號商品")
    parser.add_argument("--offline", action="store_true", help="只用快取，不補抓缺少日期")
    parser.add_argument("--csv-only", action="store_true", help="只輸出 CSV，不輸出 xlsx")
    parser.add_argument("--context-days", type=int, default=370, help="輸出起日前額外載入的日曆天數，用於 rolling 指標")
    return parser.parse_args()


def trading_weekdays(start: dt.date, end: dt.date) -> list[dt.date]:
    days: list[dt.date] = []
    current = start
    while current <= end:
        if current.weekday() < 5:
            days.append(current)
        current += dt.timedelta(days=1)
    return days


def read_history(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, dtype={"symbol": "string"})
    df["date"] = pd.to_datetime(df["date"]).dt.date
    df["symbol"] = df["symbol"].astype(str).str.strip()
    for column in ["open", "high", "low", "close", "adjusted_close", "volume", "amount"]:
        if column in df.columns:
            df[column] = pd.to_numeric(df[column], errors="coerce")
    return df


def normalize_records(day: dt.date, market: str, label: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in rows:
        values = to_template_row(market, row)
        if market == "listed":
            symbol, name = values[0], values[1]
            volume, amount = values[2], values[4]
            open_price, high, low, close = values[5], values[6], values[7], values[8]
        elif market == "mainboard":
            symbol, name = values[0], values[1]
            close, open_price, high, low = values[2], values[4], values[5], values[6]
            volume, amount = values[7], values[8]
        elif market == "esb":
            symbol, name = values[0], values[1]
            high, low, close = values[7], values[8], values[10]
            open_price = close
            volume, amount = values[13], None
        else:
            continue
        records.append(
            {
                "date": day,
                "symbol": str(symbol).strip(),
                "name": name,
                "market": label,
                "open": open_price,
                "high": high,
                "low": low,
                "close": close,
                "adjusted_close": close,
                "volume": volume,
                "amount": amount,
            }
        )
    return records


def collect_missing_history(
    missing_days: list[dt.date],
    config_path: Path,
    include_esb_latest: bool,
    include_non_stock: bool,
    request_delay: float,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    today = dt.date.today()
    for day in missing_days:
        print(f"補抓 {day.isoformat()}")
        day_records = 0
        for source in load_sources(config_path, day):
            if not source.historical and not (include_esb_latest and day == today):
                print(f"  {source.label}: skipped (latest-only source)")
                continue
            try:
                rows = source_rows(source)
            except Exception as exc:
                print(f"  {source.label}: skipped ({exc})")
                continue
            if not include_non_stock:
                rows = filter_stock_rows(rows)
            normalized = normalize_records(day, source.market, source.label, rows)
            records.extend(normalized)
            day_records += len(normalized)
            print(f"  {source.label}: {len(normalized)} rows")
            if request_delay > 0:
                time.sleep(request_delay)
        print(f"  合計: {day_records} rows")
    return pd.DataFrame(records)


def load_or_collect_history(args: argparse.Namespace, start: dt.date, end: dt.date) -> pd.DataFrame:
    cache_path = Path(args.history_cache)
    seed_path = Path(args.seed_cache) if args.seed_cache else None
    if cache_path.exists():
        history = read_history(cache_path)
    elif seed_path and seed_path.exists():
        history = read_history(seed_path)
    else:
        history = pd.DataFrame(columns=BASE_COLUMNS + ["adjusted_close"])

    cached_dates = set(history["date"].unique()) if not history.empty else set()
    context_start = start - dt.timedelta(days=max(args.context_days, 0))
    wanted_days = trading_weekdays(start, end)
    missing_days = [day for day in wanted_days if day not in cached_dates]
    if missing_days and args.offline:
        print(f"離線模式: 缺少 {len(missing_days)} 個平日資料，不補抓。")
    elif missing_days:
        fresh = collect_missing_history(
            missing_days,
            Path(args.config),
            args.include_esb_latest,
            args.include_non_stock,
            args.request_delay,
        )
        if not fresh.empty:
            history = pd.concat([history, fresh], ignore_index=True)

    if history.empty:
        raise RuntimeError("沒有可用歷史資料。")

    for column in ["open", "high", "low", "close", "adjusted_close", "volume", "amount"]:
        if column in history.columns:
            history[column] = pd.to_numeric(history[column], errors="coerce")
    history["date"] = pd.to_datetime(history["date"]).dt.date
    history["symbol"] = history["symbol"].astype(str).str.strip()
    history = (
        history.dropna(subset=["date", "symbol", "close"])
        .drop_duplicates(["date", "symbol", "market"], keep="last")
        .sort_values(["date", "market", "symbol"])
        .reset_index(drop=True)
    )
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    history.assign(date=history["date"].astype(str)).to_csv(cache_path, index=False)
    return history[(history["date"] >= context_start) & (history["date"] <= end)].copy()


def volume_status(ratio: pd.Series) -> pd.Series:
    status = pd.Series("", index=ratio.index, dtype="object")
    status = status.mask(ratio < 0.8, "縮量")
    status = status.mask((ratio >= 0.8) & (ratio <= 1.2), "正常")
    status = status.mask((ratio > 1.2) & (ratio <= 2.0), "放量")
    status = status.mask(ratio > 2.0, "爆量")
    return status


def add_indicator_columns_for_symbol(group: pd.DataFrame) -> pd.DataFrame:
    group = group.sort_values("date").copy()
    close = group["close"].astype(float)
    high = group["high"].astype(float)
    low = group["low"].astype(float)
    volume = group["volume"].astype(float)

    for window in [20, 60, 120, 240]:
        rolling_high = high.rolling(window, min_periods=window).max()
        rolling_low = low.rolling(window, min_periods=window).min()
        group[f"high_{window}d"] = rolling_high
        group[f"low_{window}d"] = rolling_low
        group[f"dist_to_{window}d_high_pct"] = (close - rolling_high) / (rolling_high + EPS) * 100
        group[f"dist_to_{window}d_low_pct"] = (close - rolling_low) / (rolling_low + EPS) * 100
        group[f"new_high_{window}d"] = (close >= rolling_high).map({True: "Yes", False: "No"})

    for window in [5, 10, 20, 60, 120, 240]:
        ma = close.rolling(window, min_periods=window).mean()
        group[f"ma_{window}"] = ma
        group[f"close_to_ma_{window}_pct"] = (close - ma) / (ma + EPS) * 100

    for window in [5, 20, 60]:
        group[f"volume_ma_{window}"] = volume.rolling(window, min_periods=window).mean()

    group["volume_ratio_20"] = volume / (group["volume_ma_20"] + EPS)
    group["volume_status"] = volume_status(group["volume_ratio_20"])
    return group


def build_indicators(history: pd.DataFrame) -> pd.DataFrame:
    parts = []
    for _, group in history.sort_values(["symbol", "date"]).groupby("symbol", sort=False):
        parts.append(add_indicator_columns_for_symbol(group))
    indicators = pd.concat(parts, ignore_index=True)
    indicators = indicators.sort_values(["date", "market", "symbol"]).reset_index(drop=True)
    output_columns = [column for column in EXPORT_COLUMNS if column in indicators.columns]
    return indicators[output_columns].rename(columns=EXPORT_COLUMNS)


def output_folder(root: Path, day: dt.date) -> Path:
    quarter = (day.month - 1) // 3 + 1
    return root / f"{day.year}" / f"Q{quarter}" / f"{day.month:02d}"


def autosize_sheet(path: Path) -> None:
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        header = str(column_cells[0].value or "")
        width = min(max(len(header) + 2, 10), 18)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            header = worksheet.cell(1, cell.column).value
            if header in NUMERIC_EXPORT_COLUMNS and isinstance(cell.value, (int, float)):
                if "%" in str(header) or header == "量比20MA":
                    cell.number_format = "0.00"
                elif "量" in str(header) or header == "成交金額":
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "0.00"
    workbook.save(path)


def write_daily_outputs(indicators: pd.DataFrame, output_root: Path, csv_only: bool) -> dict[str, Any]:
    output_root.mkdir(parents=True, exist_ok=True)
    written_files = 0
    trading_days = 0
    for day_text, day_frame in indicators.groupby("日期", sort=True):
        day = dt.date.fromisoformat(str(day_text))
        folder = output_folder(output_root, day)
        folder.mkdir(parents=True, exist_ok=True)
        base = folder / f"stock_indicators_{day.strftime('%Y%m%d')}"
        day_frame.to_csv(base.with_suffix(".csv"), index=False, encoding="utf-8-sig")
        written_files += 1
        if not csv_only:
            xlsx_path = base.with_suffix(".xlsx")
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                day_frame.to_excel(writer, sheet_name="指標", index=False)
            autosize_sheet(xlsx_path)
            written_files += 1
        trading_days += 1
    return {
        "trading_days": trading_days,
        "written_files": written_files,
        "output_root": str(output_root),
    }


def write_summary(indicators: pd.DataFrame, output_root: Path, summary: dict[str, Any]) -> Path:
    summary_path = output_root / "run_summary.json"
    payload = {
        "rows": int(len(indicators)),
        "date_start": str(indicators["日期"].min()),
        "date_end": str(indicators["日期"].max()),
        "trading_days": summary["trading_days"],
        "written_files": summary["written_files"],
        "output_root": summary["output_root"],
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "notes": [
            "不包含趨勢分數與綜合評分。",
            "興櫃官方來源目前設定為 latest-only，歷史區間預設不回補興櫃。",
            "百分比欄位以百分點表示，例如 -2.86 代表 -2.86%。",
        ],
    }
    summary_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary_path


def main() -> None:
    args = parse_args()
    start = dt.date.fromisoformat(args.start)
    end = dt.date.fromisoformat(args.end)
    if end < start:
        raise ValueError("--end 不可早於 --start")
    history = load_or_collect_history(args, start, end)
    print(f"歷史資料: rows={len(history)}, dates={history['date'].nunique()}, symbols={history['symbol'].nunique()}")
    indicators = build_indicators(history)
    indicators = indicators[
        (pd.to_datetime(indicators["日期"]).dt.date >= start)
        & (pd.to_datetime(indicators["日期"]).dt.date <= end)
    ].copy()
    output_root = Path(args.output_dir)
    summary = write_daily_outputs(indicators, output_root, args.csv_only)
    summary_path = write_summary(indicators, output_root, summary)
    print(f"指標輸出: {summary['written_files']} files, {summary['trading_days']} trading days")
    print(summary_path)


if __name__ == "__main__":
    main()
